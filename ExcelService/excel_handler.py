from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from models import ExcelTemplate, ExcelColumn, ExcelColumnType, ParsedRow, ExcelParseResult
from typing import BinaryIO
import io
from datetime import datetime
import re


class ExcelHandler:
    def __init__(self):
        pass

    def generate_template_excel(self, template: ExcelTemplate) -> bytes:
        """Template'e gore ornek Excel dosyasi olusturur"""
        wb = Workbook()
        ws = wb.active
        ws.title = template.sheet_name

        # Baslik satirini ekle (bold, mavi arka plan)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        # Kolonlari siraya gore ekle
        sorted_columns = sorted(template.columns, key=lambda x: x.order)

        for idx, column in enumerate(sorted_columns, start=1):
            cell = ws.cell(row=1, column=idx)
            cell.value = column.column_name
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

            # Ornek veri ekle (2. satir)
            sample_cell = ws.cell(row=2, column=idx)
            if column.sample_value:
                sample_cell.value = column.sample_value
            else:
                sample_cell.value = self._get_sample_value(column)

            # Kolon genisligini ayarla
            ws.column_dimensions[self._get_column_letter(idx)].width = 20

        # Aciklama satiri ekle (3. satir, italik)
        for idx, column in enumerate(sorted_columns, start=1):
            if column.description:
                desc_cell = ws.cell(row=3, column=idx)
                desc_cell.value = f"({column.description})"
                desc_cell.font = Font(italic=True, size=9, color="666666")

        # Excel'i binary olarak dondur
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    def parse_excel(self, file_content: bytes, template: ExcelTemplate) -> ExcelParseResult:
        """Excel dosyasini parse eder ve validate eder"""
        wb = load_workbook(io.BytesIO(file_content), data_only=True)
        ws = wb.active

        # Kolon mapping olustur (Excel kolon adi -> property_name)
        column_map = {}
        sorted_columns = sorted(template.columns, key=lambda x: x.order)

        for idx, column in enumerate(sorted_columns, start=1):
            column_map[idx] = column

        rows = []
        start_row = 2 if template.has_header else 1

        for row_idx, row in enumerate(ws.iter_rows(min_row=start_row, values_only=False), start=start_row):
            # Bos satirlari atla
            if all(cell.value is None or str(cell.value).strip() == "" for cell in row):
                continue

            parsed_row = ParsedRow(row_number=row_idx, data={}, errors=[], is_valid=True)

            for col_idx, cell in enumerate(row, start=1):
                if col_idx not in column_map:
                    continue

                column = column_map[col_idx]
                value = cell.value

                # Validate ve parse
                try:
                    validated_value = self._validate_and_parse(value, column, row_idx)
                    parsed_row.data[column.property_name] = validated_value
                except ValueError as e:
                    parsed_row.errors.append(f"{column.column_name}: {str(e)}")
                    parsed_row.is_valid = False

            rows.append(parsed_row)

        valid_count = sum(1 for r in rows if r.is_valid)

        return ExcelParseResult(
            total_rows=len(rows),
            valid_rows=valid_count,
            invalid_rows=len(rows) - valid_count,
            rows=rows
        )

    def _validate_and_parse(self, value: any, column: ExcelColumn, row_number: int):
        """Degeri validate eder ve gerekirse donusturur"""
        # Null kontrolu
        if value is None or str(value).strip() == "":
            if column.is_required:
                raise ValueError("Zorunlu alan bos birakilamaz")
            return None

        value_str = str(value).strip()

        # Tip bazli validasyon
        if column.column_type == ExcelColumnType.NUMBER:
            try:
                return float(value_str.replace(",", "."))
            except:
                raise ValueError("Gecerli bir sayi giriniz")

        elif column.column_type == ExcelColumnType.DATE:
            if isinstance(value, datetime):
                return value.isoformat()
            try:
                # Farkli tarih formatlarini dene
                for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%d.%m.%Y"]:
                    try:
                        dt = datetime.strptime(value_str, fmt)
                        return dt.isoformat()
                    except:
                        continue
                raise ValueError("Gecerli bir tarih giriniz (YYYY-MM-DD)")
            except:
                raise ValueError("Gecerli bir tarih giriniz")

        elif column.column_type == ExcelColumnType.BOOLEAN:
            lower_val = value_str.lower()
            if lower_val in ["true", "1", "yes", "evet", "e", "y"]:
                return True
            elif lower_val in ["false", "0", "no", "hayir", "h", "n"]:
                return False
            else:
                raise ValueError("Gecerli bir boolean deger giriniz (true/false)")

        elif column.column_type == ExcelColumnType.EMAIL:
            email_pattern = r'^[\w\.-]+@[\w\.-]+\.\w+$'
            if not re.match(email_pattern, value_str):
                raise ValueError("Gecerli bir email adresi giriniz")
            return value_str

        elif column.column_type == ExcelColumnType.PHONE:
            phone_clean = re.sub(r'[^\d+]', '', value_str)
            if len(phone_clean) < 10:
                raise ValueError("Gecerli bir telefon numarasi giriniz")
            return phone_clean

        elif column.column_type == ExcelColumnType.DROPDOWN:
            if column.dropdown_options and value_str not in column.dropdown_options:
                raise ValueError(f"Gecerli degerler: {', '.join(column.dropdown_options)}")
            return value_str

        # Text (default)
        if column.validation_rules:
            min_len = column.validation_rules.get("minLength")
            max_len = column.validation_rules.get("maxLength")

            if min_len and len(value_str) < min_len:
                raise ValueError(f"En az {min_len} karakter olmali")
            if max_len and len(value_str) > max_len:
                raise ValueError(f"En fazla {max_len} karakter olmali")

        return value_str

    def _get_sample_value(self, column: ExcelColumn) -> str:
        """Kolon tipine gore ornek deger doner"""
        samples = {
            ExcelColumnType.TEXT: "Ornek Metin",
            ExcelColumnType.NUMBER: "123.45",
            ExcelColumnType.DATE: "2025-01-15",
            ExcelColumnType.BOOLEAN: "true",
            ExcelColumnType.EMAIL: "ornek@email.com",
            ExcelColumnType.PHONE: "+905551234567",
            ExcelColumnType.DROPDOWN: column.dropdown_options[0] if column.dropdown_options else "Secenek1"
        }
        return samples.get(column.column_type, "")

    def _get_column_letter(self, col_num: int) -> str:
        """Kolon numarasini harfe cevirir (1 -> A, 27 -> AA)"""
        letter = ""
        while col_num > 0:
            col_num, remainder = divmod(col_num - 1, 26)
            letter = chr(65 + remainder) + letter
        return letter
