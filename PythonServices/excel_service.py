"""
Excel Import/Export Service for Secret Customer Evaluation System

This service handles:
1. Bulk assignment import from Excel
2. Evaluation results export to Excel
3. Dashboard data export to Excel
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
from datetime import datetime
from typing import List, Dict, Any, Optional
import json


class ExcelService:
    """Service for handling Excel import/export operations"""

    def __init__(self):
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    # ==================== IMPORT OPERATIONS ====================

    def parse_internal_assignments(self, file_path: str) -> List[Dict[str, Any]]:
        """
        Parse internal assignment Excel file

        Expected columns:
        - BranchId (GUID)
        - AssignedUserId (GUID)
        - DueDate (YYYY-MM-DD)

        Returns:
            List of assignment dictionaries
        """
        try:
            df = pd.read_excel(file_path, engine='openpyxl')

            # Validate required columns
            required_columns = ['BranchId', 'AssignedUserId', 'DueDate']
            missing_columns = [col for col in required_columns if col not in df.columns]

            if missing_columns:
                raise ValueError(f"Missing required columns: {', '.join(missing_columns)}")

            # Convert to list of dictionaries
            assignments = []
            for index, row in df.iterrows():
                try:
                    assignment = {
                        'branchId': str(row['BranchId']).strip(),
                        'assignedUserId': str(row['AssignedUserId']).strip(),
                        'dueDate': self._parse_date(row['DueDate']),
                        'externalEmail': None,
                        'externalName': None
                    }
                    assignments.append(assignment)
                except Exception as e:
                    raise ValueError(f"Error parsing row {index + 2}: {str(e)}")

            return assignments

        except Exception as e:
            raise Exception(f"Failed to parse internal assignments: {str(e)}")

    def parse_external_assignments(self, file_path: str) -> List[Dict[str, Any]]:
        """
        Parse external assignment Excel file

        Expected columns:
        - ExternalEmail (email)
        - ExternalName (string)
        - DueDate (YYYY-MM-DD)

        Returns:
            List of assignment dictionaries
        """
        try:
            df = pd.read_excel(file_path, engine='openpyxl')

            # Validate required columns
            required_columns = ['ExternalEmail', 'ExternalName', 'DueDate']
            missing_columns = [col for col in required_columns if col not in df.columns]

            if missing_columns:
                raise ValueError(f"Missing required columns: {', '.join(missing_columns)}")

            # Convert to list of dictionaries
            assignments = []
            for index, row in df.iterrows():
                try:
                    assignment = {
                        'branchId': None,
                        'assignedUserId': None,
                        'externalEmail': str(row['ExternalEmail']).strip(),
                        'externalName': str(row['ExternalName']).strip(),
                        'dueDate': self._parse_date(row['DueDate'])
                    }

                    # Validate email
                    if '@' not in assignment['externalEmail']:
                        raise ValueError(f"Invalid email: {assignment['externalEmail']}")

                    assignments.append(assignment)
                except Exception as e:
                    raise ValueError(f"Error parsing row {index + 2}: {str(e)}")

            return assignments

        except Exception as e:
            raise Exception(f"Failed to parse external assignments: {str(e)}")

    # ==================== EXPORT OPERATIONS ====================

    def export_evaluation_results(
        self,
        evaluations: List[Dict[str, Any]],
        output_path: str,
        project_name: str = "Evaluation Results"
    ) -> str:
        """
        Export evaluation results to Excel

        Args:
            evaluations: List of evaluation dictionaries
            output_path: Path to save the Excel file
            project_name: Name of the project

        Returns:
            Path to the created Excel file
        """
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Evaluations"

            # Add title
            ws.merge_cells('A1:H1')
            title_cell = ws['A1']
            title_cell.value = f"{project_name} - Evaluation Results"
            title_cell.font = Font(size=14, bold=True)
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 30

            # Add metadata
            ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws['A2'].font = Font(italic=True, size=9)

            # Headers
            headers = [
                'Evaluation ID',
                'Branch',
                'Evaluator',
                'Score',
                'Max Score',
                'Percentage',
                'Completed Date',
                'Status'
            ]

            row = 4
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.fill = self.header_fill
                cell.font = self.header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = self.border

            # Data rows
            for evaluation in evaluations:
                row += 1
                ws.cell(row=row, column=1, value=evaluation.get('id', ''))
                ws.cell(row=row, column=2, value=evaluation.get('branchName', 'N/A'))
                ws.cell(row=row, column=3, value=evaluation.get('evaluatorName', 'External'))
                ws.cell(row=row, column=4, value=evaluation.get('totalScore', 0))
                ws.cell(row=row, column=5, value=evaluation.get('maxScore', 0))

                # Calculate percentage
                percentage = 0
                if evaluation.get('maxScore', 0) > 0:
                    percentage = (evaluation.get('totalScore', 0) / evaluation.get('maxScore', 0)) * 100

                ws.cell(row=row, column=6, value=f"{percentage:.2f}%")
                ws.cell(row=row, column=7, value=evaluation.get('completedAt', ''))
                ws.cell(row=row, column=8, value='Completed' if evaluation.get('isCompleted') else 'Pending')

                # Apply borders
                for col in range(1, 9):
                    ws.cell(row=row, column=col).border = self.border

            # Auto-adjust column widths
            for col in range(1, 9):
                ws.column_dimensions[get_column_letter(col)].width = 18

            # Save workbook
            wb.save(output_path)
            return output_path

        except Exception as e:
            raise Exception(f"Failed to export evaluation results: {str(e)}")

    def export_dashboard_data(
        self,
        dashboard_data: Dict[str, Any],
        output_path: str
    ) -> str:
        """
        Export dashboard statistics to Excel with multiple sheets

        Args:
            dashboard_data: Dictionary containing dashboard statistics
            output_path: Path to save the Excel file

        Returns:
            Path to the created Excel file
        """
        try:
            wb = openpyxl.Workbook()

            # Remove default sheet
            wb.remove(wb.active)

            # Sheet 1: Overview
            ws_overview = wb.create_sheet("Overview")
            self._create_overview_sheet(ws_overview, dashboard_data)

            # Sheet 2: Branch Performance
            if 'branchPerformance' in dashboard_data:
                ws_branches = wb.create_sheet("Branch Performance")
                self._create_branch_performance_sheet(ws_branches, dashboard_data['branchPerformance'])

            # Sheet 3: Evaluator Performance
            if 'evaluatorPerformance' in dashboard_data:
                ws_evaluators = wb.create_sheet("Evaluator Performance")
                self._create_evaluator_performance_sheet(ws_evaluators, dashboard_data['evaluatorPerformance'])

            # Sheet 4: Trend Data
            if 'trendData' in dashboard_data:
                ws_trend = wb.create_sheet("Trend Analysis")
                self._create_trend_sheet(ws_trend, dashboard_data['trendData'])

            # Save workbook
            wb.save(output_path)
            return output_path

        except Exception as e:
            raise Exception(f"Failed to export dashboard data: {str(e)}")

    def export_checklist_template(
        self,
        checklist: Dict[str, Any],
        output_path: str
    ) -> str:
        """
        Export a checklist as an Excel template for offline evaluation

        Args:
            checklist: Checklist dictionary with sections and questions
            output_path: Path to save the Excel file

        Returns:
            Path to the created Excel file
        """
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Checklist"

            # Title
            ws.merge_cells('A1:D1')
            title_cell = ws['A1']
            title_cell.value = checklist.get('name', 'Checklist Template')
            title_cell.font = Font(size=16, bold=True)
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 40

            row = 3

            # Iterate through sections
            for section in checklist.get('sections', []):
                # Section header
                ws.merge_cells(f'A{row}:D{row}')
                section_cell = ws[f'A{row}']
                section_cell.value = section.get('title', 'Untitled Section')
                section_cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                section_cell.font = Font(bold=True, size=12)
                section_cell.alignment = Alignment(horizontal='left', vertical='center')
                ws.row_dimensions[row].height = 25
                row += 1

                # Questions
                for question in section.get('questions', []):
                    ws.cell(row=row, column=1, value=question.get('text', ''))
                    ws.cell(row=row, column=2, value=question.get('type', ''))
                    ws.cell(row=row, column=3, value=question.get('weight', 0))
                    ws.cell(row=row, column=4, value='')  # Answer column

                    # Apply borders
                    for col in range(1, 5):
                        ws.cell(row=row, column=col).border = self.border

                    row += 1

                row += 1  # Empty row between sections

            # Column headers
            ws.insert_rows(2)
            headers = ['Question', 'Type', 'Weight', 'Answer']
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=2, column=col)
                cell.value = header
                cell.fill = self.header_fill
                cell.font = self.header_font
                cell.alignment = Alignment(horizontal='center')

            # Adjust column widths
            ws.column_dimensions['A'].width = 50
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 10
            ws.column_dimensions['D'].width = 30

            wb.save(output_path)
            return output_path

        except Exception as e:
            raise Exception(f"Failed to export checklist template: {str(e)}")

    # ==================== HELPER METHODS ====================

    def _parse_date(self, date_value: Any) -> str:
        """Parse date to ISO format string"""
        if isinstance(date_value, datetime):
            return date_value.strftime('%Y-%m-%d')
        elif isinstance(date_value, str):
            # Try to parse string date
            try:
                parsed_date = pd.to_datetime(date_value)
                return parsed_date.strftime('%Y-%m-%d')
            except:
                return date_value
        else:
            return str(date_value)

    def _create_overview_sheet(self, ws, data: Dict[str, Any]):
        """Create overview summary sheet"""
        ws['A1'] = "Dashboard Overview"
        ws['A1'].font = Font(size=14, bold=True)

        row = 3
        metrics = [
            ('Total Projects', data.get('totalProjects', 0)),
            ('Active Projects', data.get('activeProjects', 0)),
            ('Total Assignments', data.get('totalAssignments', 0)),
            ('Completed Evaluations', data.get('completedEvaluations', 0)),
            ('Pending Evaluations', data.get('pendingEvaluations', 0)),
            ('Average Score', f"{data.get('averageScore', 0):.2f}%"),
            ('Total Branches', data.get('totalBranches', 0)),
            ('Total Evaluators', data.get('totalEvaluators', 0))
        ]

        for metric, value in metrics:
            ws.cell(row=row, column=1, value=metric).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            row += 1

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20

    def _create_branch_performance_sheet(self, ws, branches: List[Dict[str, Any]]):
        """Create branch performance comparison sheet"""
        # Headers
        headers = ['Branch Name', 'Total Evaluations', 'Average Score', 'Max Score', 'Min Score', 'Status']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font

        # Data
        for idx, branch in enumerate(branches, start=2):
            ws.cell(row=idx, column=1, value=branch.get('branchName', ''))
            ws.cell(row=idx, column=2, value=branch.get('totalEvaluations', 0))
            ws.cell(row=idx, column=3, value=f"{branch.get('averageScore', 0):.2f}%")
            ws.cell(row=idx, column=4, value=f"{branch.get('maxScore', 0):.2f}%")
            ws.cell(row=idx, column=5, value=f"{branch.get('minScore', 0):.2f}%")
            ws.cell(row=idx, column=6, value=branch.get('status', 'Active'))

        # Auto-adjust widths
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 18

    def _create_evaluator_performance_sheet(self, ws, evaluators: List[Dict[str, Any]]):
        """Create evaluator performance sheet"""
        # Headers
        headers = ['Evaluator Name', 'Completed', 'Pending', 'Total Assigned', 'Completion Rate']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font

        # Data
        for idx, evaluator in enumerate(evaluators, start=2):
            completed = evaluator.get('completedCount', 0)
            total = evaluator.get('totalAssigned', 0)
            completion_rate = (completed / total * 100) if total > 0 else 0

            ws.cell(row=idx, column=1, value=evaluator.get('evaluatorName', ''))
            ws.cell(row=idx, column=2, value=completed)
            ws.cell(row=idx, column=3, value=evaluator.get('pendingCount', 0))
            ws.cell(row=idx, column=4, value=total)
            ws.cell(row=idx, column=5, value=f"{completion_rate:.2f}%")

        # Auto-adjust widths
        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 20

    def _create_trend_sheet(self, ws, trend_data: List[Dict[str, Any]]):
        """Create trend analysis sheet"""
        # Headers
        headers = ['Period', 'Total Evaluations', 'Average Score', 'Completion Rate']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font

        # Data
        for idx, period in enumerate(trend_data, start=2):
            ws.cell(row=idx, column=1, value=period.get('period', ''))
            ws.cell(row=idx, column=2, value=period.get('totalEvaluations', 0))
            ws.cell(row=idx, column=3, value=f"{period.get('averageScore', 0):.2f}%")
            ws.cell(row=idx, column=4, value=f"{period.get('completionRate', 0):.2f}%")

        # Auto-adjust widths
        for col in range(1, 5):
            ws.column_dimensions[get_column_letter(col)].width = 20


# ==================== STANDALONE USAGE ====================

if __name__ == "__main__":
    import sys

    service = ExcelService()

    if len(sys.argv) < 3:
        print("Usage: python excel_service.py <command> <file_path> [additional_args]")
        print("\nCommands:")
        print("  parse_internal <file_path>")
        print("  parse_external <file_path>")
        print("  export_evaluations <data_json> <output_path>")
        print("  export_dashboard <data_json> <output_path>")
        sys.exit(1)

    command = sys.argv[1]
    file_path = sys.argv[2]

    try:
        if command == "parse_internal":
            result = service.parse_internal_assignments(file_path)
            print(json.dumps(result, indent=2))

        elif command == "parse_external":
            result = service.parse_external_assignments(file_path)
            print(json.dumps(result, indent=2))

        elif command == "export_evaluations":
            with open(file_path, 'r') as f:
                data = json.load(f)
            output_path = sys.argv[3] if len(sys.argv) > 3 else "evaluations.xlsx"
            result = service.export_evaluation_results(data, output_path)
            print(f"Exported to: {result}")

        elif command == "export_dashboard":
            with open(file_path, 'r') as f:
                data = json.load(f)
            output_path = sys.argv[3] if len(sys.argv) > 3 else "dashboard.xlsx"
            result = service.export_dashboard_data(data, output_path)
            print(f"Exported to: {result}")

        else:
            print(f"Unknown command: {command}")
            sys.exit(1)

    except Exception as e:
        print(f"Error: {str(e)}", file=sys.stderr)
        sys.exit(1)
